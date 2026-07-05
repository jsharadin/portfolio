"""
Build Chilk_Inventory_Planner.xlsx from the data/ folder.

Rerun order:  generate_sales_history.py -> forecast_demand.py -> build_workbook.py
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = ROOT / "Chilk_Inventory_Planner.xlsx"

# ---- palette ----
BROWN = "5B3A29"; CREAM = "FFF6E9"; TAN = "E8D9C4"; WHITE = "FFFFFF"
BLUE_IN = Font(name="Arial", color="0000FF")           # inputs
GREEN_LN = Font(name="Arial", color="008000")          # cross-sheet links
BODY = Font(name="Arial", size=10)
HDR = Font(name="Arial", size=10, bold=True, color=WHITE)
TITLE = Font(name="Arial", size=16, bold=True, color=BROWN)
SUB = Font(name="Arial", size=10, italic=True, color="7A6A58")
HFILL = PatternFill("solid", start_color=BROWN)
SFILL = PatternFill("solid", start_color=TAN)
THIN = Side(style="thin", color="C9B79C")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center")

FLAVORS = ["Strawberry", "Chocolate", "Cinnamon", "Banana", "Thai Tea", "Taro", "Coffee"]
PRICE = 4.99


def style_header(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = HDR, HFILL, CTR, BOX


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_title(ws, text, sub, ncols):
    ws["A1"] = text; ws["A1"].font = TITLE
    ws["A2"] = sub; ws["A2"].font = SUB


def main():
    ing = pd.read_csv(DATA / "ingredients.csv")
    weekly = pd.read_csv(DATA / "weekly_demand.csv", parse_dates=["week_start"])
    sales_raw = pd.read_csv(DATA / "sales_history.csv", parse_dates=["week_start"])
    backtest = pd.read_csv(DATA / "backtest.csv")
    fc_raw = pd.read_csv(DATA / "forecast.csv")
    fc_sum = fc_raw[fc_raw.week_start == "SUMMARY"].set_index("flavor")
    fc_wk = fc_raw[fc_raw.week_start != "SUMMARY"].copy()
    fc_wk["week_start"] = pd.to_datetime(fc_wk["week_start"])
    fc_pivot = fc_wk.pivot(index="week_start", columns="flavor", values="forecast_bottles")[FLAVORS]

    wb = Workbook()

    # ================= README =================
    ws = wb.active; ws.title = "README"; ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 110])
    ws["B2"] = "CHILK — Inventory Planning & Demand Forecasting Model"; ws["B2"].font = Font(name="Arial", size=18, bold=True, color=BROWN)
    ws["B3"] = "Rebuild of the operational inventory tracker I built and ran as COO of Chilk (plant-based milk startup, 2021–2022)."; ws["B3"].font = SUB
    lines = [
        ("", ""),
        ("ABOUT", "Chilk produced bottled oat-milk drinks in 7 flavors, sold at LA/SB/Berkeley farmers markets, Duffl delivery, and campus partners. "
                  "This workbook converts a demand forecast into ingredient purchase orders, accounting for pack sizes, current stock, supplier lead times, and safety stock."),
        ("", ""),
        ("DATA NOTE", "Sales history is SIMULATED (seeded, reproducible) but calibrated to real Chilk records: real channels, flavors, prices, observed farmers-market volumes, "
                      "and academic-calendar seasonality. All ingredient data (recipes, pack sizes, costs, suppliers, lead times) is real, from Chilk's 2022 operating files."),
        ("", ""),
        ("HOW TO USE", ""),
        ("  1.", "Review the DASHBOARD for current forecast, order cost, and alerts."),
        ("  2.", "On PRODUCTION PLAN, set the production date and adjust planned bottles per flavor (blue cells). Defaults come from the 4-week forecast + safety stock."),
        ("  3.", "ORDER PLAN computes what to buy: net requirements, packs to order, cost, and order-by dates. Red = order is already late."),
        ("  4.", "FORECAST and SALES HISTORY show the demand model behind the defaults (Holt damped-trend exponential smoothing, 8-week horizon)."),
        ("  5.", "To refresh with new data, rerun the pipeline: generate_sales_history.py → forecast_demand.py → build_workbook.py."),
        ("", ""),
        ("COLOR KEY", "Blue = input you can change   |   Black = formula   |   Green = link to another sheet   |   Red/yellow highlight = action needed"),
    ]
    r = 5
    for label, text in lines:
        ws.cell(row=r, column=2, value=(label + "  " + text).strip() if label and text else (label or text))
        cell = ws.cell(row=r, column=2)
        cell.font = Font(name="Arial", size=10, bold=label.isupper() and not text) if not text else BODY
        if label.isupper() and text:
            cell.value = f"{label}:  {text}"
            cell.font = BODY
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30 if len(str(cell.value or "")) > 90 else 16
        r += 1

    # ================= Sales History =================
    ws = wb.create_sheet("Sales History"); ws.sheet_view.showGridLines = False
    sheet_title(ws, "Weekly Bottles Sold by Flavor", "Simulated history calibrated to real Chilk channel volumes — see README", 10)
    hdr_row = 4
    heads = ["Week"] + FLAVORS + ["Total", "Revenue ($)"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header(ws, hdr_row, 1, len(heads))
    for i, (_, row) in enumerate(weekly.iterrows()):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=1, value=row["week_start"].date()).number_format = "mm/dd/yy"
        for j, f in enumerate(FLAVORS, 2):
            ws.cell(row=r, column=j, value=int(row[f]))
        ws.cell(row=r, column=9, value=f"=SUM(B{r}:H{r})")
        ws.cell(row=r, column=10, value=f"=I{r}*{PRICE}").number_format = "$#,##0"
        for c in range(1, 11):
            ws.cell(row=r, column=c).font = BODY; ws.cell(row=r, column=c).border = BOX
    last_hist = hdr_row + len(weekly)
    tr = last_hist + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
    for c in range(2, 11):
        col = get_column_letter(c)
        ws.cell(row=tr, column=c, value=f"=SUM({col}{hdr_row+1}:{col}{last_hist})")
        ws.cell(row=tr, column=c).font = Font(name="Arial", bold=True)
        ws.cell(row=tr, column=c).fill = SFILL; ws.cell(row=tr, column=c).border = BOX
    ws.cell(row=tr, column=10).number_format = "$#,##0"
    set_widths(ws, [10] + [10] * 7 + [10, 12])
    ws.freeze_panes = "B5"

    chart = LineChart(); chart.title = "Weekly Bottles Sold (All Channels)"; chart.style = 12
    chart.height, chart.width = 8, 24; chart.y_axis.title = "Bottles"; chart.legend = None
    chart.add_data(Reference(ws, min_col=9, min_row=hdr_row, max_row=last_hist), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=hdr_row + 1, max_row=last_hist))
    ws.add_chart(chart, "L4")

    # ================= Forecast =================
    ws = wb.create_sheet("Forecast"); ws.sheet_view.showGridLines = False
    sheet_title(ws, "8-Week Demand Forecast", "Holt damped-trend exponential smoothing per flavor (pipeline/forecast_demand.py)", 9)
    hdr_row = 4
    heads = ["Week"] + FLAVORS + ["Total"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header(ws, hdr_row, 1, len(heads))
    for i, (wk, row) in enumerate(fc_pivot.iterrows()):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=1, value=wk.date()).number_format = "mm/dd/yy"
        for j, f in enumerate(FLAVORS, 2):
            ws.cell(row=r, column=j, value=float(row[f])).number_format = "0.0"
        ws.cell(row=r, column=9, value=f"=SUM(B{r}:H{r})").number_format = "0.0"
        for c in range(1, 10):
            ws.cell(row=r, column=c).font = BODY; ws.cell(row=r, column=c).border = BOX
    fc_last = hdr_row + len(fc_pivot)

    # summary block
    sr = fc_last + 3
    ws.cell(row=sr, column=1, value="Forecast Summary (per flavor)").font = Font(name="Arial", size=12, bold=True, color=BROWN)
    heads = ["Flavor", "Avg Sold (last 4 wks)", "Forecast (next 4 wks)", "Safety Stock (bottles)"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=sr + 1, column=c, value=h)
    style_header(ws, sr + 1, 1, 4)
    for i, f in enumerate(FLAVORS):
        r = sr + 2 + i
        ws.cell(row=r, column=1, value=f)
        ws.cell(row=r, column=2, value=float(fc_sum.loc[f, "avg_last_4wk"])).number_format = "0.0"
        # next-4-week forecast as live formula over the weekly block above
        col = get_column_letter(2 + i)
        ws.cell(row=r, column=3, value=f"=SUM({col}{hdr_row+1}:{col}{hdr_row+4})").number_format = "0.0"
        ws.cell(row=r, column=4, value=int(fc_sum.loc[f, "safety_stock"]))
        for c in range(1, 5):
            ws.cell(row=r, column=c).font = BODY; ws.cell(row=r, column=c).border = BOX
    set_widths(ws, [10, 19, 19, 20, 10, 10, 10, 10, 10])

    chart = LineChart(); chart.title = "Forecast by Flavor (bottles/week)"; chart.style = 12
    chart.height, chart.width = 9, 18; chart.y_axis.title = "Bottles"
    chart.add_data(Reference(ws, min_col=2, max_col=8, min_row=hdr_row, max_row=fc_last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=hdr_row + 1, max_row=fc_last))
    ws.add_chart(chart, "K4")

    # backtest accuracy block
    br = sr + 2 + len(FLAVORS) + 2
    ws.cell(row=br, column=1, value="Model Accuracy — 8-Week Holdout Backtest").font = Font(name="Arial", size=12, bold=True, color=BROWN)
    ws.cell(row=br + 1, column=1, value="Model trained on all weeks except the last 8, then compared to what actually sold.").font = SUB
    heads = ["Flavor", "Actual (8 wks)", "Predicted (8 wks)", "Abs % Error"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=br + 2, column=c, value=h)
    style_header(ws, br + 2, 1, 4)
    bt_first = br + 3
    for i, (_, g) in enumerate(backtest.iterrows()):
        r = bt_first + i
        ws.cell(row=r, column=1, value=g.flavor)
        ws.cell(row=r, column=2, value=int(g.actual_8wk))
        ws.cell(row=r, column=3, value=float(g.predicted_8wk)).number_format = "0.0"
        ws.cell(row=r, column=4, value=f"=ABS(C{r}-B{r})/B{r}").number_format = "0.0%"
        for c in range(1, 5):
            ws.cell(row=r, column=c).font = BODY; ws.cell(row=r, column=c).border = BOX
    bt_last = bt_first + len(backtest) - 1
    r = bt_last + 1
    ws.cell(row=r, column=1, value="MAPE (mean)").font = Font(name="Arial", bold=True)
    ws.cell(row=r, column=4, value=f"=AVERAGE(D{bt_first}:D{bt_last})").number_format = "0.0%"
    ws.cell(row=r, column=4).font = Font(name="Arial", bold=True)
    for c in range(1, 5):
        ws.cell(row=r, column=c).fill = SFILL; ws.cell(row=r, column=c).border = BOX

    # ================= Sales Data (raw) =================
    ws = wb.create_sheet("Sales Data")
    heads = ["week_start", "channel", "flavor", "forecasted", "delivered", "sold", "unsold", "revenue"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 1, len(heads))
    for i, (_, g) in enumerate(sales_raw.iterrows()):
        r = 2 + i
        ws.cell(row=r, column=1, value=g.week_start.date()).number_format = "mm/dd/yy"
        ws.cell(row=r, column=2, value=g.channel)
        ws.cell(row=r, column=3, value=g.flavor)
        for c, k in [(4, "forecasted"), (5, "delivered"), (6, "sold"), (7, "unsold")]:
            ws.cell(row=r, column=c, value=int(g[k]))
        ws.cell(row=r, column=8, value=float(g.revenue)).number_format = "#,##0.00"
        for c in range(1, 9):
            ws.cell(row=r, column=c).font = BODY
    raw_last = 1 + len(sales_raw)
    set_widths(ws, [11, 24, 12, 11, 10, 8, 8, 10])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{raw_last}"

    # ================= Channel Insights =================
    ws = wb.create_sheet("Channel Insights"); ws.sheet_view.showGridLines = False
    sheet_title(ws, "Channel Performance", "All metrics computed live from the Sales Data sheet with SUMIFS", 8)
    channels = sorted(sales_raw.channel.unique())
    hdr_row = 4
    heads = ["Channel", "Bottles Sold", "Revenue ($)", "Rev Share", "Delivered", "Unsold", "Waste Rate", "Sell-Through"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header(ws, hdr_row, 1, len(heads))
    ch_first = hdr_row + 1
    for i, ch in enumerate(channels):
        r = ch_first + i
        ws.cell(row=r, column=1, value=ch)
        ws.cell(row=r, column=2, value=f"=SUMIFS('Sales Data'!$F$2:$F${raw_last},'Sales Data'!$B$2:$B${raw_last},A{r})").number_format = "#,##0"
        ws.cell(row=r, column=3, value=f"=SUMIFS('Sales Data'!$H$2:$H${raw_last},'Sales Data'!$B$2:$B${raw_last},A{r})").number_format = "$#,##0"
        ws.cell(row=r, column=4, value=f"=C{r}/$C${ch_first+len(channels)}").number_format = "0.0%"
        ws.cell(row=r, column=5, value=f"=SUMIFS('Sales Data'!$E$2:$E${raw_last},'Sales Data'!$B$2:$B${raw_last},A{r})").number_format = "#,##0"
        ws.cell(row=r, column=6, value=f"=SUMIFS('Sales Data'!$G$2:$G${raw_last},'Sales Data'!$B$2:$B${raw_last},A{r})").number_format = "#,##0"
        ws.cell(row=r, column=7, value=f"=F{r}/E{r}").number_format = "0.0%"
        ws.cell(row=r, column=8, value=f"=B{r}/E{r}").number_format = "0.0%"
        for c in range(1, 9):
            ws.cell(row=r, column=c).font = BODY; ws.cell(row=r, column=c).border = BOX
    ch_total = ch_first + len(channels)
    ws.cell(row=ch_total, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
    for c, fmt in [(2, "#,##0"), (3, "$#,##0"), (5, "#,##0"), (6, "#,##0")]:
        col = get_column_letter(c)
        ws.cell(row=ch_total, column=c, value=f"=SUM({col}{ch_first}:{col}{ch_total-1})").number_format = fmt
    ws.cell(row=ch_total, column=4, value=1).number_format = "0%"
    ws.cell(row=ch_total, column=7, value=f"=F{ch_total}/E{ch_total}").number_format = "0.0%"
    ws.cell(row=ch_total, column=8, value=f"=B{ch_total}/E{ch_total}").number_format = "0.0%"
    for c in range(1, 9):
        ws.cell(row=ch_total, column=c).font = Font(name="Arial", bold=True)
        ws.cell(row=ch_total, column=c).fill = SFILL; ws.cell(row=ch_total, column=c).border = BOX
    set_widths(ws, [26, 12, 12, 10, 11, 9, 11, 13])

    bar = BarChart(); bar.type = "bar"; bar.title = "Revenue by Channel ($)"; bar.style = 10; bar.legend = None
    bar.height, bar.width = 9, 13
    bar.add_data(Reference(ws, min_col=3, min_row=ch_first, max_row=ch_total - 1))
    bar.set_categories(Reference(ws, min_col=1, min_row=ch_first, max_row=ch_total - 1))
    s = bar.series[0]; s.graphicalProperties.solidFill = BROWN
    ws.add_chart(bar, "J4")

    bar = BarChart(); bar.title = "Waste Rate by Channel"; bar.style = 10; bar.legend = None
    bar.height, bar.width = 9, 13; bar.y_axis.numFmt = "0%"
    bar.add_data(Reference(ws, min_col=7, min_row=ch_first, max_row=ch_total - 1))
    bar.set_categories(Reference(ws, min_col=1, min_row=ch_first, max_row=ch_total - 1))
    s = bar.series[0]; s.graphicalProperties.solidFill = "B08968"
    ws.add_chart(bar, "J23")

    # ================= Production Plan =================
    ws = wb.create_sheet("Production Plan"); ws.sheet_view.showGridLines = False
    sheet_title(ws, "Production Plan", "Blue cells are inputs — adjust planned bottles, everything downstream recalculates", 6)
    ws["A4"] = "Production date:"; ws["A4"].font = Font(name="Arial", bold=True)
    ws["B4"] = "=TODAY()+7"; ws["B4"].font = BLUE_IN; ws["B4"].number_format = "mm/dd/yyyy"
    ws["C4"] = "← default: 1 week out (Chilk produced the day before events). Overwrite with a fixed date if needed."; ws["C4"].font = SUB
    hdr_row = 6
    heads = ["Flavor", "Forecast Next 4 Wks", "Safety Stock", "Planned Bottles", "% of Plan", "Batches (24/btl)"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header(ws, hdr_row, 1, 6)
    n = len(FLAVORS)
    for i, f in enumerate(FLAVORS):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=1, value=f).font = BODY
        ws.cell(row=r, column=2, value=f"=Forecast!C{fc_last+5+i}").font = GREEN_LN
        ws.cell(row=r, column=2).number_format = "0.0"
        ws.cell(row=r, column=3, value=f"=Forecast!D{fc_last+5+i}").font = GREEN_LN
        ws.cell(row=r, column=4, value=f"=ROUNDUP(B{r}+C{r},0)")  # default plan; user may overwrite -> becomes input
        ws.cell(row=r, column=4).font = BLUE_IN
        ws.cell(row=r, column=5, value=f"=D{r}/$D${hdr_row+n+1}").number_format = "0.0%"
        ws.cell(row=r, column=6, value=f"=D{r}/24").number_format = "0.0"
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BOX
    tr = hdr_row + n + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
    for c, fmt in [(2, "0.0"), (3, "0"), (4, "0"), (6, "0.0")]:
        col = get_column_letter(c)
        ws.cell(row=tr, column=c, value=f"=SUM({col}{hdr_row+1}:{col}{hdr_row+n})")
        ws.cell(row=tr, column=c).number_format = fmt
    ws.cell(row=tr, column=5, value=1).number_format = "0%"
    for c in range(1, 7):
        ws.cell(row=tr, column=c).font = Font(name="Arial", bold=True)
        ws.cell(row=tr, column=c).fill = SFILL; ws.cell(row=tr, column=c).border = BOX
    ws.cell(row=tr + 2, column=1, value="Projected revenue @ $4.99/bottle:").font = Font(name="Arial", bold=True)
    ws.cell(row=tr + 2, column=4, value=f"=D{tr}*{PRICE}").number_format = "$#,##0"
    set_widths(ws, [14, 19, 13, 15, 10, 15])

    pp_first, pp_last, pp_total = hdr_row + 1, hdr_row + n, tr

    # ================= Order Plan =================
    ws = wb.create_sheet("Order Plan"); ws.sheet_view.showGridLines = False
    sheet_title(ws, "Purchase Order Plan", "Driven by Production Plan — net requirements, packs to order, cost, and order-by dates", 17)
    hdr_row = 4
    heads = ["Ingredient", "Category", "Applies To", "Unit", "Per Bottle", "Bottles Applied",
             "Required", "On Hand", "Net Required", "Pack Size", "Packs to Order", "Order Qty",
             "Pack Cost ($)", "Order Cost ($)", "Supplier", "Lead Time (days)", "Order By", "Status"]
    for c, h in enumerate(heads, 1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header(ws, hdr_row, 1, len(heads))
    for i, (_, g) in enumerate(ing.iterrows()):
        r = hdr_row + 1 + i
        v = [g.ingredient, g.category, g.applies_to, g.unit, g.per_bottle]
        for c, val in enumerate(v, 1):
            ws.cell(row=r, column=c, value=val)
        # bottles applied: ALL -> total planned, else that flavor's planned bottles
        ws.cell(row=r, column=6,
                value=f"=IF(C{r}=\"ALL\",'Production Plan'!$D${pp_total},"
                      f"SUMIF('Production Plan'!$A${pp_first}:$A${pp_last},C{r},'Production Plan'!$D${pp_first}:$D${pp_last}))")
        ws.cell(row=r, column=7, value=f"=E{r}*F{r}").number_format = "#,##0.0"
        ws.cell(row=r, column=8, value=float(g.on_hand)).font = BLUE_IN
        ws.cell(row=r, column=8).number_format = "#,##0.0"
        ws.cell(row=r, column=9, value=f"=MAX(0,G{r}-H{r})").number_format = "#,##0.0"
        ws.cell(row=r, column=10, value=float(g.pack_size))
        ws.cell(row=r, column=11, value=f"=IF(I{r}=0,0,ROUNDUP(I{r}/J{r},0))")
        ws.cell(row=r, column=12, value=f"=K{r}*J{r}").number_format = "#,##0"
        ws.cell(row=r, column=13, value=float(g.pack_cost)).number_format = "#,##0.00"
        ws.cell(row=r, column=14, value=f"=K{r}*M{r}").number_format = "$#,##0"
        ws.cell(row=r, column=15, value=g.supplier)
        ws.cell(row=r, column=16, value=int(g.lead_time_days))
        ws.cell(row=r, column=17, value=f"=IF(K{r}=0,\"—\",'Production Plan'!$B$4-P{r})").number_format = "mm/dd/yy"
        ws.cell(row=r, column=18,
                value=f"=IF(K{r}=0,\"OK\",IF('Production Plan'!$B$4-P{r}<TODAY(),\"LATE\",\"ORDER\"))")
        for c in range(1, 19):
            cell = ws.cell(row=r, column=c)
            if cell.font == Font():
                cell.font = BODY
            if not cell.font.color or cell.font.color.rgb in (None, "FF000000"):
                cell.font = BODY if c != 8 else BLUE_IN
            cell.border = BOX
        ws.cell(row=r, column=18).alignment = CTR
    op_last = hdr_row + len(ing)
    tr = op_last + 1
    ws.cell(row=tr, column=1, value="TOTAL ORDER COST").font = Font(name="Arial", bold=True)
    ws.cell(row=tr, column=14, value=f"=SUM(N{hdr_row+1}:N{op_last})").number_format = "$#,##0"
    ws.cell(row=tr, column=14).font = Font(name="Arial", bold=True)
    for c in range(1, 19):
        ws.cell(row=tr, column=c).fill = SFILL; ws.cell(row=tr, column=c).border = BOX
    rng = f"R{hdr_row+1}:R{op_last}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"LATE"'], fill=PatternFill("solid", start_color="F4CCCC")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"ORDER"'], fill=PatternFill("solid", start_color="FFF2CC")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", start_color="D9EAD3")))
    set_widths(ws, [22, 12, 11, 6, 10, 13, 11, 10, 12, 10, 12, 10, 11, 12, 24, 12, 11, 9])
    ws.freeze_panes = "B5"

    # ================= Dashboard =================
    ws = wb.create_sheet("Dashboard", 1); ws.sheet_view.showGridLines = False
    for c in range(1, 15):
        ws.cell(row=1, column=c).fill = HFILL
        ws.cell(row=2, column=c).fill = HFILL
        ws.cell(row=3, column=c).fill = HFILL
    ws["B2"] = "CHILK — Inventory & Ordering Dashboard"; ws["B2"].font = Font(name="Arial", size=18, bold=True, color=WHITE)
    ws["B3"] = '=CONCATENATE("Production date: ",TEXT(\'Production Plan\'!B4,"mmm dd, yyyy"))'
    ws["B3"].font = Font(name="Arial", size=10, italic=True, color=TAN)
    kpis = [
        ("Planned Bottles", f"='Production Plan'!D{pp_total}", "#,##0"),
        ("Projected Revenue", f"='Production Plan'!D{pp_total}*{PRICE}", "$#,##0"),
        ("Total Order Cost", f"='Order Plan'!N{op_last+1}", "$#,##0"),
        ("Items to Order", f"=COUNTIF('Order Plan'!R5:R{op_last},\"ORDER\")+COUNTIF('Order Plan'!R5:R{op_last},\"LATE\")", "0"),
        ("Late Orders", f"=COUNTIF('Order Plan'!R5:R{op_last},\"LATE\")", "0"),
        ("Hist. Waste Rate", f"='Channel Insights'!G{ch_total}", "0.0%"),
    ]
    col = 2
    for label, formula, fmt in kpis:
        ws.cell(row=5, column=col, value=label).font = Font(name="Arial", size=9, bold=True, color="7A6A58")
        c = ws.cell(row=6, column=col, value=formula)
        c.font = Font(name="Arial", size=20, bold=True, color=BROWN); c.number_format = fmt
        for rr in (5, 6):
            ws.cell(row=rr, column=col).fill = PatternFill("solid", start_color=CREAM)
            ws.cell(row=rr, column=col).border = BOX
            ws.cell(row=rr, column=col).alignment = CTR
        ws.column_dimensions[get_column_letter(col)].width = 17
        col += 2
    ws.column_dimensions["A"].width = 2
    for c in range(3, 14, 2):
        ws.column_dimensions[get_column_letter(c)].width = 2

    # chart: planned bottles by flavor
    bar = BarChart(); bar.title = "Planned Production by Flavor"; bar.style = 10; bar.legend = None
    bar.height, bar.width = 8, 11; bar.y_axis.title = "Bottles"
    bar.add_data(Reference(wb["Production Plan"], min_col=4, min_row=pp_first, max_row=pp_last))
    bar.set_categories(Reference(wb["Production Plan"], min_col=1, min_row=pp_first, max_row=pp_last))
    bar.series[0].graphicalProperties.solidFill = BROWN
    ws.add_chart(bar, "B9")

    # chart: order cost by ingredient
    bar2 = BarChart(); bar2.type = "bar"; bar2.title = "Order Cost by Ingredient ($)"; bar2.style = 11; bar2.legend = None
    bar2.height, bar2.width = 8, 11
    bar2.add_data(Reference(wb["Order Plan"], min_col=14, min_row=5, max_row=op_last))
    bar2.set_categories(Reference(wb["Order Plan"], min_col=1, min_row=5, max_row=op_last))
    bar2.series[0].graphicalProperties.solidFill = "B08968"
    ws.add_chart(bar2, "H9")

    ws["B26"] = "Workflow:  Sales History → Channel Insights → Forecast → Production Plan (adjust blue cells) → Order Plan (buy list) — see README"
    ws["B26"].font = SUB

    # sheet display order
    order = ["README", "Dashboard", "Sales History", "Channel Insights", "Forecast",
             "Production Plan", "Order Plan", "Sales Data"]
    wb._sheets = [wb[n] for n in order]

    for s in wb.worksheets:
        for row in s.iter_rows():
            for cell in row:
                if cell.value is not None and cell.font.name != "Arial":
                    f = cell.font
                    cell.font = Font(name="Arial", size=f.size or 10, bold=f.bold, italic=f.italic, color=f.color)

    wb.save(OUT)
    print(f"saved {OUT}")


if __name__ == "__main__":
    main()
